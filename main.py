import datetime
import math
import multiprocessing
import os
import sys
from sys import platform
import time
import csv
import random
from multiprocessing import Process
from random import randint
from time import localtime, strftime, time
import pandas as pd
import psutil
import xlsxwriter
import random
import openpyxl
from openpyxl.chart.axis import DateAxis

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, Reference, ScatterChart
from openpyxl.chart.series import SeriesLabel, Series

from PySide6.QtGui import QColor
from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog
from PySide6 import QtGui, QtWidgets, QtCore
from PySide6.QtCore import QTimer, QThread
from PySide6.QtCore import QFile
from PySide6.QtGui import QPainter
from PySide6.QtCharts import QChart, QChartView
from ui_main import Ui_MainWindow
from scipy.interpolate import make_interp_spline
import numpy as np
from chart import Chart
import pyqtgraph as pg
from usb_connections import communication
from usb_connections.communication import sensor_unit, sensor_name, sensor_val
from usb_connections.msc import get_device_path, write_bin_file_to_device

QtGui.QImageReader.setAllocationLimit(0)

shadow_elements = {
    "frame",
    "frame_2",
}

wai = [0xBD, 0xBA, 0xBE,
       0xBC, 0xBF, 0xBB,
       0xB1, 0xB2, 0xB3]


class MscInfo(QThread):
    def __init__(self, usb: communication.DlmmUSB):
        super().__init__()
        self.__usb = usb

    def disk(self):
        d = psutil.disk_partitions()

    def run(self):

        try:
            p = Process(target=psutil.disk_partitions)
            p.start()
            # p.join()
            while p.is_alive():
                QThread.sleep(1)
            self.__usb.get_msc_info()
        except TypeError as e:
            QThread.sleep(5)
            # print(f"MSC not loaded {e}, sleep 10 sec.")
            p = Process(target=psutil.disk_partitions)
            p.start()
            # p.join()
            while p.is_alive():
                # print("procc is alive")
                QThread.sleep(1)

            self.__usb.get_msc_info()


class Downloader(QThread):
    def __init__(self, usb: communication.DlmmUSB, sensors: list[communication.SensorSd], index: int):
        super().__init__()
        self._usb = usb
        self._sensors = sensors
        self._index = index

    def run(self):
        # self._usb.get_sd_data(self._sensors, self._index)
        self._usb.get_msc_data(self._sensors, self._index)


class DownloaderApp(QThread):
    def __init__(self, usb: communication.DlmmUSB, dev_path, filename, content, old_filename):
        super().__init__()
        self.__usb = usb
        self.__dev_path = dev_path
        self.__filename = filename
        self.__content = content
        self.__old_filename = old_filename

    def run(self):
        self.__usb.load_new_app(self.__dev_path, self.__filename, self.__content, self.__old_filename)


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        # self.graphWidget = pg.GraphicsLayoutWidget()
        # graphWidget sd
        self.time0 = 0
        self.downloader = None
        self.downloader_app = None
        self.msc_info = None
        self.graphWidget_sd = pg.PlotWidget()
        self.x_plot_sd = list(range(1000))
        self.y_plot_sd = list(range(1000))
        self.t_plot_sd = list(range(1000))
        self.graphWidget_sd.setBackground('w')
        self.pen_sd = pg.mkPen(color=(255, 0, 0), width=2)
        # self.data_line = self.graphWidget.addPlot(self.x_plot, self.y_plot, pen=self.pen)
        self.data_line_sd = self.graphWidget_sd.plot(self.x_plot_sd, self.y_plot_sd, pen=self.pen_sd)
        # self.graphWidget.addItem(self.label)
        self.graphWidget_sd.showGrid(x=True, y=True, alpha=1.0)
        self.vLine_sd = pg.InfiniteLine(angle=90, movable=False)
        self.hLine_sd = pg.InfiniteLine(angle=0, movable=False)
        self.graphWidget_sd.addItem(self.vLine_sd, ignoreBounds=True)
        self.graphWidget_sd.addItem(self.hLine_sd, ignoreBounds=True)
        # self.graphWidget.setAutoPan(y=True, x=True)
        self.vb_sd = self.graphWidget_sd.plotItem.vb
        self.graphWidget_sd.scene().sigMouseMoved.connect(self.mouseMoved_sd)
        self.ui.gridLayout_3.addWidget(self.graphWidget_sd)
        ######################################

        self.graphWidget = pg.PlotWidget()
        self.x_plot = list(range(1000))
        self.y_plot = [0 for _ in range(1000)]
        self.graphWidget.setBackground('w')
        # self.label = pg.LabelItem(text='test', justify='right')

        self.pen = pg.mkPen(color=(255, 0, 0), width=2)
        # self.data_line = self.graphWidget.addPlot(self.x_plot, self.y_plot, pen=self.pen)
        self.data_line = self.graphWidget.plot(self.x_plot, self.y_plot, pen=self.pen)
        # self.graphWidget.addItem(self.label)
        self.graphWidget.showGrid(x=True, y=True, alpha=1.0)
        self.vLine = pg.InfiniteLine(angle=90, movable=False)
        self.hLine = pg.InfiniteLine(angle=0, movable=False)
        self.graphWidget.addItem(self.vLine, ignoreBounds=True)
        self.graphWidget.addItem(self.hLine, ignoreBounds=True)
        # self.graphWidget.setAutoPan(y=True, x=True)
        self.vb = self.graphWidget.plotItem.vb
        self.graphWidget.scene().sigMouseMoved.connect(self.mouseMoved)
        self.ui.gridLayout.addWidget(self.graphWidget)
        self.clock_timer = QTimer()
        self.clock_timer.setInterval(500)
        self.clock_timer.timeout.connect(self.clock)
        self.clock_timer.start()

        # self.usb = communication.UsbCommunication()

        self.dl = communication.DlmmUSB(0x52A4, 0x0483)
        self.connection_timer = QTimer()
        self.connection_timer.setInterval(500)
        self.connection_timer.timeout.connect(self.check_device_connection)
        self.connection_timer.start()
        self.names_of_sensor = ["Датчик шума", "Термопара", "Датчик тока",
                                "Датчик пульса", "Датчик света", "Датчик ультрафиолета",
                                "Датчик давления", "Датчик влажности", "Датчик температуры"]
        self.sensor_btns = [self.ui.btn_conn_1, self.ui.btn_conn_2, self.ui.btn_conn_3, self.ui.btn_conn_4,
                            self.ui.btn_conn_5, self.ui.btn_conn_6, self.ui.btn_conn_7, self.ui.btn_conn_8,
                            self.ui.btn_conn_9, self.ui.btn_conn_10, self.ui.btn_conn_11]
        self.sensor_names = [self.ui.label_name_1, self.ui.label_name_2, self.ui.label_name_3, self.ui.label_name_4,
                             self.ui.label_name_5, self.ui.label_name_6, self.ui.label_name_7, self.ui.label_name_8,
                             self.ui.label_name_9, self.ui.label_name_10, self.ui.label_name_11]
        self.sensor_values = [self.ui.label_value_1, self.ui.label_value_2, self.ui.label_value_3,
                              self.ui.label_value_4,
                              self.ui.label_value_5, self.ui.label_value_6, self.ui.label_value_7,
                              self.ui.label_value_8,
                              self.ui.label_value_9, self.ui.label_value_10, self.ui.label_value_11]
        self.sensor_btns_start = [self.ui.btn_start_1, self.ui.btn_start_2, self.ui.btn_start_3,
                                  self.ui.btn_start_4, self.ui.btn_start_5, self.ui.btn_start_6,
                                  self.ui.btn_start_7, self.ui.btn_start_8, self.ui.btn_start_9,
                                  self.ui.btn_start_10, self.ui.btn_start_11]
        self.sensor_btns_stop = [self.ui.btn_stop_1, self.ui.btn_stop_2, self.ui.btn_stop_3,
                                 self.ui.btn_stop_4, self.ui.btn_stop_5, self.ui.btn_stop_6,
                                 self.ui.btn_stop_7, self.ui.btn_stop_8, self.ui.btn_stop_9,
                                 self.ui.btn_stop_10, self.ui.btn_stop_11]

        self.sensor_edits_hz = [self.ui.lineEdit_hz_1, self.ui.lineEdit_hz_2, self.ui.lineEdit_hz_3,
                                self.ui.lineEdit_hz_4, self.ui.lineEdit_hz_5, self.ui.lineEdit_hz_6,
                                self.ui.lineEdit_hz_7, self.ui.lineEdit_hz_8, self.ui.lineEdit_hz_9,
                                self.ui.lineEdit_hz_10, self.ui.lineEdit_hz_11]
        self.sensor_frames = [self.ui.frame_sens_1, self.ui.frame_sens_2, self.ui.frame_sens_3,
                              self.ui.frame_sens_4, self.ui.frame_sens_5, self.ui.frame_sens_6,
                              self.ui.frame_sens_7, self.ui.frame_sens_8, self.ui.frame_sens_9,
                              self.ui.frame_sens_10, self.ui.frame_sens_11]
        self.btns_sd = [self.ui.btn_sd_1, self.ui.btn_sd_2, self.ui.btn_sd_3, self.ui.btn_sd_4,
                        self.ui.btn_sd_5, self.ui.btn_sd_6, self.ui.btn_sd_7, self.ui.btn_sd_8,
                        self.ui.btn_sd_9, self.ui.btn_sd_10, self.ui.btn_sd_11]
        # self.ui.btn_sd_10.setVisible(False)
        # self.ui.btn_sd_11.setVisible(False)
        self.last_btn_clicked = 0
        self.ui.btn_start.clicked.connect(self.btn_start_action)
        self.ui.btn_stop.clicked.connect(self.btn_stop_action)
        self.ui.lineEdit.textChanged.connect(self.edit_hz_one_action)
        self.ui.comboBox.activated.connect(self.comboactivated)
        self.ui.btn_sd_clear.clicked.connect(self.sd_clear_action)
        for i in self.btns_sd:
            i.clicked.connect(self.btns_sd_action)
        for i in self.sensor_edits_hz:
            i.textChanged.connect(self.edit_hz_action)
        for i in self.sensor_btns_start:
            i.clicked.connect(self.btn_start_one_sensor)
        for i in self.sensor_btns_stop:
            i.clicked.connect(self.btn_stop_one_sensor)

        self.timers = [QTimer() for _ in range(11)]
        for i in self.timers:
            i.setTimerType(QtCore.Qt.TimerType.PreciseTimer)
            i.timeout.connect(self.handle_timeout)

        self.sensors = [communication.Sensor(index=i) for i in range(11)]
        self.sensors_sd = [communication.SensorSd(index=i) for i in range(11)]
        for i in self.timers:
            i.setInterval(1000)
        self.device_connected = 1
        self.check_device_connection()
        # self.sensors_id = self.usb.check_connections()
        self.set_shadow()
        self.check_connections()
        self.ui.btn_sd_data.clicked.connect(self.btn_sd_data_action)
        self.ui.btn_all_sensors.clicked.connect(self.btn_all_sensors_action)
        self.ui.btn_sd_download.clicked.connect(self.btn_download_action)
        self.ui.btn_sd_save.clicked.connect(self.btn_save_action)
        self.ui.btn_save.clicked.connect(self.btn_save_main_action)
        self.ui.btn_clear.clicked.connect(self.btn_clear_action)
        self.ui.btn_load_app.clicked.connect(self.btn_load_action)
        for i in self.sensor_btns:
            i.clicked.connect(self.btn_sensor_action)
        now = strftime("%Y", localtime())
        self.ui.label_5.setText(f'© DCIE {now}')
        self.wai_sd = 0x0

    def btn_load_action(self):

        path = get_device_path()


        if path is None:
            return
        file = QFileDialog.getOpenFileName(self)

        if file[0][-4:] != ".bin":
            # print(file[0][-4:])

            self.ui.btn_load_app.setText("Файл повреждён")
            return
        filename = file[0][-12:]
        with open(file[0], "rb") as f:
            # print(file[0])
            file_content = f.read()
            #for i in file_content:
               # self.dl.load_app.append(i)
        #delta_len_app = 212992 - len(self.dl.load_app)
        delta_len_app = 212992 - len(file_content)
        if delta_len_app < 0:
            self.ui.btn_load_app.setText("Файл повреждён")
            return
        #write_bin_file_to_device(path, file[0][-12:], file_content)
        # self.dl.load_app += [0xFF for _ in range(delta_len_app)]
        self.ui.btn_load_app.setText("Загрузка")
        self.downloader_app = DownloaderApp(self.dl, path[0], filename, file_content, path[1])
        self.downloader_app.finished.connect(self.download_app_finished)
        self.downloader_app.start()

    def download_app_finished(self):
        self.ui.btn_load_app.setText("Перезагрузите устройство")
        del self.downloader_app

    def export(self, dataset: dict, file_path: str):
        try:
            if dataset['Время']:
                df = pd.DataFrame(dataset)
                df.drop(labels=[0, 1, 2, 3], axis=0, inplace=True)
                df.reset_index(drop=True, inplace=True)
                # print(df)
                # df.to_excel(writer, sheet_name=sensor_name[self.sensors[self.last_btn_clicked].who_am_i])
                wb = openpyxl.Workbook()
                ws = wb.active

                redFill = PatternFill(start_color='f0f0f0',
                                      end_color='f0f0f0',
                                      fill_type='solid')
                blackFill = PatternFill(start_color='00808080',
                                        end_color='00808080',
                                        fill_type='solid')
                thin = Side(border_style="thin", color="000000")
                # wb.create_sheet(sensor_name[self.sensors[self.last_btn_clicked].who_am_i], 0)
                ws.title = sensor_name[self.sensors[self.last_btn_clicked].who_am_i]
                for i in dataframe_to_rows(df, index=True, header=True):
                    ws.append(i)

                ch = ws['B2']
                ws.freeze_panes = ch
                min_row = ws.min_row
                max_row = ws.max_row
                min_col = ws.min_column
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 20
                ch1 = ws.cell(1, min_col)
                ch2 = ws.cell(1, min_col + 1)
                ch3 = ws.cell(1, min_col + 2)
                ch4 = ws.cell(1, min_col + 3)
                ch1.fill = blackFill
                ch2.fill = blackFill
                ch3.fill = blackFill
                ch4.fill = blackFill
                for i in range(1, max_row):
                    c = ws.cell(i + 1, min_col)
                    c1 = ws.cell(i + 1, min_col + 1)
                    c2 = ws.cell(i + 1, min_col + 2)
                    c3 = ws.cell(i + 1, min_col + 3)
                    c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    c1.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    c2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    c3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c1.alignment = Alignment(horizontal="center", vertical="center")
                    c2.alignment = Alignment(horizontal="center", vertical="center")
                    c3.alignment = Alignment(horizontal="center", vertical="center")
                    if i % 2 == 0:
                        c.fill = redFill
                        c1.fill = redFill
                        c2.fill = redFill
                        c3.fill = redFill
                ws.delete_rows(min_row + 1)
                ws.delete_rows(min_row + 1)

                chart = ScatterChart()
                chart.height = 15
                chart.width = 20
                # data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=max_row)
                data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=max_row)
                chart.title = sensor_name[self.sensors[self.last_btn_clicked].who_am_i]
                chart.style = 13
                chart.y_axis.title = sensor_unit[self.sensors[self.last_btn_clicked].who_am_i]
                # chart.x_axis = DateAxis(crossAx=100)
                # chart.x_axis.number_format = 'M-ss'
                # chart.x_axis.majorTimeUnit = "days"
                chart.x_axis.title = "Время эксперимента, с"
                # chart.add_data(data, titles_from_data=True)
                dates = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=max_row)
                series = openpyxl.chart.Series(data, dates, title_from_data=True)
                # chart.set_categories(dates)
                chart.series.append(series)
                chart.legend.title = sensor_unit[self.sensors[self.last_btn_clicked].who_am_i]
                ws.add_chart(chart, "F2")

                wb.save(file_path)
        except FileNotFoundError as e:
            print(f'File save error: {e}')

    def export_sd(self, sensors: list, file_path: str):
        wb = openpyxl.Workbook()
        # ws = wb.active
        # ws = []
        count = 0
        '''for i, j in enumerate(sensors):
            ws.append(wb.create_sheet(j.name, i))'''
        try:
            for i, j in enumerate(sensors):

                if j.dataset['Время']:
                    ws = wb.create_sheet(j.name, count)
                    count += 1
                    df = pd.DataFrame(j.dataset)
                    # df.to_excel(writer, sheet_name=sensor_name[self.sensors[self.last_btn_clicked].who_am_i])

                    redFill = PatternFill(start_color='f0f0f0',
                                          end_color='f0f0f0',
                                          fill_type='solid')
                    blackFill = PatternFill(start_color='00808080',
                                            end_color='00808080',
                                            fill_type='solid')
                    thin = Side(border_style="thin", color="000000")
                    # wb.create_sheet(sensor_name[self.sensors[self.last_btn_clicked].who_am_i], 0)
                    ws.title = sensor_name[j.who_am_i]
                    for k in dataframe_to_rows(df, index=True, header=True):
                        ws.append(k)

                    ch = ws['B2']
                    ws.freeze_panes = ch
                    min_row = ws.min_row
                    max_row = ws.max_row
                    min_col = ws.min_column
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 25
                    ws.column_dimensions['D'].width = 20
                    ch1 = ws.cell(1, min_col)
                    ch2 = ws.cell(1, min_col + 1)
                    ch3 = ws.cell(1, min_col + 2)
                    ch4 = ws.cell(1, min_col + 3)
                    ch1.fill = blackFill
                    ch2.fill = blackFill
                    ch3.fill = blackFill
                    ch4.fill = blackFill
                    for l in range(1, max_row):
                        c = ws.cell(l + 1, min_col)
                        c1 = ws.cell(l + 1, min_col + 1)
                        c2 = ws.cell(l + 1, min_col + 2)
                        c3 = ws.cell(l + 1, min_col + 3)
                        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        c1.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        c2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        c3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        c.alignment = Alignment(horizontal="center", vertical="center")
                        c1.alignment = Alignment(horizontal="center", vertical="center")
                        c2.alignment = Alignment(horizontal="center", vertical="center")
                        c3.alignment = Alignment(horizontal="center", vertical="center")
                        if l % 2 == 0:
                            c.fill = redFill
                            c1.fill = redFill
                            c2.fill = redFill
                            c3.fill = redFill
                    ws.delete_rows(min_row + 1)
                    ws.delete_rows(min_row + 1)

                    chart = LineChart()
                    chart.height = 15
                    chart.width = 20
                    # data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=max_row)
                    data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=max_row)
                    chart.title = sensor_name[j.who_am_i]
                    chart.style = 11
                    chart.y_axis.title = sensor_unit[j.who_am_i]
                    # chart.x_axis = DateAxis(crossAx=100)
                    # chart.x_axis.number_format = 'M-ss'
                    # chart.x_axis.majorTimeUnit = "days"
                    chart.x_axis.title = "Время эксперимента, с"
                    chart.add_data(data, titles_from_data=True)
                    dates = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=max_row)
                    # series = openpyxl.chart.Series(data, dates, title_from_data=True)
                    chart.set_categories(dates)
                    # chart.series.append(series)
                    chart.legend.title = sensor_unit[j.who_am_i]
                    ws.add_chart(chart, "F2")

            wb.save(file_path)
        except FileNotFoundError as e:
            print(f'File save error: {e}')

    def btn_clear_action(self):
        self.sensors[self.last_btn_clicked].x = list(range(100))
        self.sensors[self.last_btn_clicked].y = [0 for _ in range(100)]
        self.sensors[self.last_btn_clicked].x_time = list()
        self.sensors[self.last_btn_clicked].x_timestamp = list()
        self.x_plot = list(range(1000))
        self.y_plot = [0 for _ in range(1000)]
        self.data_line.setData(self.x_plot, self.y_plot)
        self.graphWidget.setXRange(int(self.x_plot[0]), int(self.x_plot[-1]), padding=None, update=True)

    def btn_save_main_action(self):
        if self.timers[self.last_btn_clicked].isActive():
            self.timers[self.last_btn_clicked].stop()
            self.sensors[self.last_btn_clicked].running = False
        dataset = {'Время': self.sensors[self.last_btn_clicked].x_time,
                   'Время эксперимента': self.sensors[self.last_btn_clicked].x_timestamp,
                   f'Значение, {sensor_unit[self.sensors[self.last_btn_clicked].who_am_i]}':
                       self.sensors[self.last_btn_clicked].y[100:]}

        file_path = QFileDialog.getSaveFileName(self, 'Сохранить')[0]
        if file_path[-5:] != '.xlsx':
            file_path += '.xlsx'
        self.export(dataset, file_path)
        if not self.timers[self.last_btn_clicked].isActive():
            self.timers[self.last_btn_clicked].start()
            self.sensors[self.last_btn_clicked].running = True

    def btn_save_action(self):
        title = 'Сохранить'

        file_path = QFileDialog.getSaveFileName(self, title)[0]
        if file_path[-5:] != '.xlsx':
            file_path += '.xlsx'
        self.export_sd(self.sensors_sd, file_path)
        '''try:
            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                for i in self.sensors_sd:
                    if i.dataset['Время']:
                        # print(i.dataset)
                        df = pd.DataFrame(i.dataset)
                        df.to_excel(writer, sheet_name=i.name)
        except FileNotFoundError as e:
            print(f'File save error: {e}')'''

    def sd_clear_action(self):
        self.dl.clear_sd()
        self.btn_sd_data_action()

    def btns_sd_action(self):
        sender = self.sender()
        idx = 0
        for i, j in enumerate(self.btns_sd):
            if sender is j:
                idx = i
        self.wai_sd = self.sensors_sd[idx].who_am_i
        self.x_plot_sd = np.linspace(min(self.sensors_sd[idx].x), max(self.sensors_sd[idx].x),
                                     len(self.sensors_sd[idx].x) * 10)
        self.t_plot_sd = self.sensors_sd[idx].dataset['Время']
        spl = make_interp_spline(self.sensors_sd[idx].x, self.sensors_sd[idx].y, 3)

        self.y_plot_sd = spl(self.x_plot_sd)
        self.data_line_sd.setData(self.x_plot_sd, self.y_plot_sd)
        self.graphWidget_sd.setXRange(int(self.x_plot_sd[0]), int(self.x_plot_sd[-1]), padding=None, update=True)
        name = sensor_val[self.sensors_sd[idx].who_am_i]
        unit = sensor_unit[self.sensors_sd[idx].who_am_i]
        self.graphWidget_sd.setLabel('left', f'{name}, {unit}')
        self.graphWidget_sd.setLabel('bottom', "X")

    def btn_download_action(self):
        for idx in range(len(self.timers)):
            if self.timers[idx].isActive():
                self.timers[idx].stop()
                self.sensors[idx].running = False
        for i in self.btns_sd:
            i.setEnabled(False)
        for i in self.sensors_sd:
            i.x.clear()
            i.y.clear()
        index = self.ui.comboBox.currentIndex()
        self.downloader = Downloader(self.dl, self.sensors_sd, index)
        self.downloader.finished.connect(self.download_finished)
        self.downloader.start()
        # self.dl.get_sd_data(self.sensors_sd)

    def download_finished(self):
        self.ui.btn_sd_download.setEnabled(False)
        self.ui.btn_sd_save.setEnabled(True)
        for i, j in enumerate(self.sensors_sd):
            if j.available:
                # self.btns_sd[i].setVisible(True)
                self.btns_sd[i].setEnabled(True)
                self.btns_sd[i].setText(j.name)
        del self.downloader

    def comboactivated(self, index):
        self.ui.label_sd_detect.setText(f'Выбран {index + 1} файл')
        self.ui.label_sd_date.setText(self.sender().itemText(index))
        self.ui.label_sd_size.setText(f'{self.dl.sd_files[index].size} кб')
        self.ui.btn_sd_download.setEnabled(True)

    def btn_sd_data_action(self):
        if not self.dl.device:
            self.ui.label_sd_detect.setText("Карта памяти не обнаружена")
            self.ui.btn_sd_clear.setEnabled(False)
        self.ui.btn_sd_download.setEnabled(False)
        self.ui.btn_sd_save.setEnabled(False)
        if not self.dl.data_downloaded:
            for i in self.btns_sd:
                i.setEnabled(False)
                # i.setVisible(False)
        # self.dl.get_sd_info()
        self.ui.stackedWidget.setCurrentWidget(self.ui.sd_page)
        # print("set sd_page")
        self.msc_info = MscInfo(self.dl)
        self.msc_info.finished.connect(self.msc_info_finished)
        self.msc_info.start()
        # self.dl.get_msc_info()
        # print(self.ui.comboBox.currentIndex())

    def msc_info_finished(self):
        for i in range(self.ui.comboBox.count()):
            # print(i)
            self.ui.comboBox.removeItem(self.ui.comboBox.currentIndex())
        for i, j in enumerate(self.dl.sd_files):
            time = j.date.strftime('%Y-%m-%d %H:%M:%S')
            self.ui.comboBox.addItem(f'{time}')
        if self.dl.sd_empty != 1:
            files = len(self.dl.sd_files)
            if files == 4 or files == 1:
                self.ui.label_sd_detect.setText(f'На карте {len(self.dl.sd_files)} файл')
            else:
                self.ui.label_sd_detect.setText(f'На карте {len(self.dl.sd_files)} файла')
            self.ui.label_sd_date.setText('Выберите файл')
        if self.dl.sd_last_data is not None:
            date_sd = self.dl.sd_last_data.strftime('%Y-%m-%d')
            self.ui.label_sd_date.setText('Последняя запись ' + date_sd)
        self.ui.label_sd_size.setText(f'{self.dl.sd_last_data_size} кбайт')
        del self.msc_info


    def clock(self):
        # print(localtime())
        curr_time = strftime("%H:%M:%S %Y-%m-%d", localtime())
        self.ui.label_16.setText(curr_time)

    def mouseMoved_sd(self, evt):

        pos = evt
        mousePoint = self.vb_sd.mapSceneToView(pos)
        index = mousePoint.x()
        # print(index, self.x_plot_sd[-1], len(self.x_plot_sd), len(self.y_plot_sd))
        # y_ind = round((index - self.x_plot_sd[0]) * 10) + 10
        y_ind = int(index) * 10 + 10
        if y_ind >= len(self.x_plot_sd):
            y_ind = len(self.x_plot_sd) - 1
        elif y_ind <= 0:
            y_ind = 0
        y_val = round(self.y_plot_sd[y_ind], 2)

        try:
            if int(index) >= 0:
                idx = 0
                for i, j in enumerate(self.sensors_sd):
                    if j.who_am_i == self.wai_sd:
                        idx = i
                t = str(self.sensors_sd[idx].dataset["Время эксперимента"][int(index)])
                # print(t)
                # t = self.t_plot_sd[int(index)]
            else:
                # t = "-" * 11 + "---"
                t = "-" * 5

        except IndexError:
            # t = "-" * 11 + "---"
            t = "-" * 5
        except KeyError:
            t = "-" * 5
        if type(t) is not str:
            # t = "-" * 11 + "---"
            t = "-" * 5
        if self.wai_sd != 0:
            name = sensor_val[self.wai_sd]
            unit = sensor_unit[self.wai_sd]
        else:
            name = "y"
            unit = ""
        # print(t)
        self.graphWidget_sd.setTitle(
            f"<span style='font-size: 14pt; color: green'>t={t},&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span>"
            f"<span style='font-size: 14pt; color: red'>{name}={y_val} {unit},&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span>"
            f"<span style='font-size: 11pt; color: gray'>x={round(index, 2)}</span>")
        self.vLine_sd.setPos(mousePoint.x())
        self.hLine_sd.setPos(mousePoint.y())

    def mouseMoved(self, evt):
        pos = evt
        mousePoint = self.vb.mapSceneToView(pos)
        index = mousePoint.x()
        if self.ui.checkBox.isChecked():
            ind = int((index - self.x_plot[0]) * 10) + 8
            if ind > 999:
                ind = 999
            if ind < 0:
                ind = 0
        else:
            ind = int(index - self.x_plot[0])
            if ind > 99:
                ind = 99
            if ind < 0:
                ind = 0
        '''y_ind = round((index - self.x_plot[0]) * 10) + 10
        if y_ind >= 1000:
            y_ind = 999
        elif y_ind <= 0:
            y_ind = 0'''
        # print(len(self.sensors[self.last_btn_clicked].x_time), index, ind)
        try:
            t = self.sensors[self.last_btn_clicked].x_timestamp[int(index - 100)]
        except IndexError:
            t = "-" * 5
        # print(t)
        y_val = round(self.y_plot[ind], 2)
        wai = self.sensors[self.last_btn_clicked].who_am_i
        name = sensor_val[wai]
        unit = sensor_unit[wai]
        self.graphWidget.setTitle(
            f"<span style='font-size: 14pt; color: green'>t={t},&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span>"
            f"<span style='font-size: 14pt; color: red'>{name}={y_val} {unit},&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span>"
            f"<span style='font-size: 11pt; color: gray'>x={round(index, 2)}</span>")
        self.vLine.setPos(mousePoint.x())
        self.hLine.setPos(mousePoint.y())

    @staticmethod
    def example_func(x):
        return math.sin(x * 3) + 4 * math.sin(x)

    def update_data(self, i: int):

        # self.sensors[i].x = self.sensors[i].x[1:]
        self.sensors[i].x.append(self.sensors[i].x[-1] + 1)
        t = strftime("%Y-%m-%d %H:%M:%S", localtime())
        ts = datetime.datetime.utcnow().strftime('.%f')[:-3]
        t += ts
        self.sensors[i].x_time.append(t)
        time0 = datetime.datetime.strptime(self.sensors[i].x_time[0], "%Y-%m-%d %H:%M:%S.%f")
        time1 = datetime.datetime.strptime(t, "%Y-%m-%d %H:%M:%S.%f")

        timestamp = round((time1 - time0).total_seconds(), 2)
        self.sensors[i].x_timestamp.append(timestamp)
        self.ui.btn_clear.setEnabled(True)
        data = self.dl.get_data_from_sensor(self.sensors[i])
        self.sensors[i].y.append(round(data + random.uniform(0.01, 0.05), 2))
        if self.last_btn_clicked != i:
            return
        if self.ui.checkBox.isChecked():
            self.x_plot = np.linspace(min(self.sensors[i].x[-100:]), max(self.sensors[i].x[-100:]), 1000)
            spl = make_interp_spline(self.sensors[i].x[-100:], self.sensors[i].y[-100:])

            self.y_plot = spl(self.x_plot)
        else:
            self.x_plot = self.sensors[i].x[-100:]

            self.y_plot = self.sensors[i].y[-100:]
        for i, j in enumerate(self.y_plot):
            if j < 0:
                self.y_plot[i] = 0

    def edit_hz_one_action(self):
        sender = self.sender()
        try:
            # print(f"new hz: {sender.text()}")
            new_hz = int(sender.text())

        except ValueError:
            new_hz = 1
        if new_hz > 1000:
            new_hz = 1000
            sender.setText(str(new_hz))
        self.sensors[self.last_btn_clicked].frequency = new_hz
        interval = int(1 / new_hz * 1000)
        if interval == 0:
            interval = 1
        self.timers[self.last_btn_clicked].setInterval(interval)

    def edit_hz_action(self):
        sender = self.sender()
        idx = 0
        for i, j in enumerate(self.sensor_edits_hz):
            if sender is j:
                idx = i
        try:
            new_hz = int(sender.text())
        except ValueError:
            new_hz = 1
        if new_hz > 1000:
            new_hz = 1000
            sender.setText(str(new_hz))
        self.sensors[idx].frequency = new_hz
        interval = int(1 / new_hz * 1000)
        if interval == 0:
            interval = 1
        self.timers[idx].setInterval(interval)

    def btn_start_one_sensor(self):
        # print(self.sender)
        sender = self.sender()
        idx = 0
        for i, j in enumerate(self.sensor_btns_start):
            if sender is j:
                idx = i
                break
        if not self.timers[idx].isActive():
            self.timers[idx].start()
            self.sensors[idx].running = True
            self.sensors[idx].timestamp = round(time(), 2)

    def btn_stop_one_sensor(self):

        sender = self.sender()
        idx = 0
        for i, j in enumerate(self.sensor_btns_stop):
            if sender is j:
                idx = i
                break
        if self.timers[idx].isActive():
            self.timers[idx].stop()
            self.sensors[idx].running = False

    def handle_timeout(self):
        # print(time() - self.time0)

        sender = self.sender()
        idx = 0
        for i, j in enumerate(self.timers):
            if j is sender:
                idx = i
                break

        if self.sensors[idx].running:
            self.update_data(idx)

        if self.last_btn_clicked == idx:
            # print(self.x_plot[0])
            # print(self.sensors[idx].unit)
            self.ui.label_sensor_value.setText(f'{round(self.sensors[idx].y[-1], 2)} {self.sensors[idx].unit}')
            # self.data_line.setData(self.sensors[idx].x, self.sensors[idx].y)
            self.data_line.setData(self.x_plot, self.y_plot)
            self.graphWidget.setXRange(int(self.x_plot[0]), int(self.x_plot[-1]), padding=None, update=True)
        self.sensor_values[idx].setText(f'{round(self.sensors[idx].y[-1], 2)} {self.sensors[idx].unit}')
        # print(f'handle timeout time: {time() - self.time0}')
        self.time0 = time()

    def check_device_connection(self):
        # print('check_device_connection')
        self.dl.get_available_sensors(self.sensors)
        self.check_connections()
        if self.dl.first_connection:
            self.dl.set_date()
            self.dl.first_connection = False
        # print(self.dl.device)
        if self.dl.device is not None:
            self.ui.label_15.setStyleSheet("""background-color: #00AE68""")
            self.ui.label_6.setText("Модуль подключен")

        else:
            self.ui.label_15.setStyleSheet("""background-color: #bf0600""")
            self.ui.label_6.setText("Модуль отключен")

    def btn_all_sensors_action(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.all_sensors)
        for i, j in enumerate(self.sensor_edits_hz):
            # print(self.sensors[i].frequency)
            j.setText(str(self.sensors[i].frequency))

    def btn_sensor_action(self):
        # print(self.sender().text())
        sender = 0
        for i in self.names_of_sensor:
            # print(j, self.sender().text())
            for k, j in enumerate(self.sensors):
                if j.name == self.sender().text():
                    sender = k
        # print(self.sender().text())
        # sender = int(self.sender().text().split()[0]) - 1
        # print(sender)
        self.dl.send_who_am_i(self.sensors[sender].who_am_i)
        if not self.sensors[sender].x_time:
            self.ui.btn_clear.setEnabled(False)
        self.last_btn_clicked = sender
        self.ui.stackedWidget.setCurrentWidget(self.ui.one_sensor)
        # print(self.last_btn_clicked, self.sensors[sender].y)
        self.ui.label_sensor_name.setText(self.sensors[sender].name)
        self.ui.lineEdit.setText(str(self.sensors[self.last_btn_clicked].frequency))
        self.data_line.setData(self.sensors[sender].x, self.sensors[sender].y)
        name = sensor_val[self.sensors[sender].who_am_i]
        unit = sensor_unit[self.sensors[sender].who_am_i]
        self.graphWidget.setLabel('left', f'{name}, {unit}')
        self.graphWidget.setLabel('bottom', "X")

        # self.data_line.setData(self.x_plot, self.y_plot)
        if self.timers[sender].isActive():

            self.ui.label_sensor_value.setText(f'{round(self.sensors[sender].y[-1], 2)}')

        else:
            self.ui.label_sensor_value.setText("------")
        # print(f"btn {self.sender().text()} checked")

    def btn_start_action(self):
        # print(self.last_btn_clicked)
        if not self.timers[self.last_btn_clicked].isActive():
            self.timers[self.last_btn_clicked].start()
            self.sensors[self.last_btn_clicked].running = True
            self.sensors[self.last_btn_clicked].timestamp = round(time(), 2)

    def btn_stop_action(self):
        if self.timers[self.last_btn_clicked].isActive():
            self.timers[self.last_btn_clicked].stop()
            self.ui.label_sensor_value.setText("------")
            self.sensor_values[self.last_btn_clicked].setText("------")
            self.sensors[self.last_btn_clicked].running = False

    def check_connections(self):
        # self.ui.frame_15.setVisible(True)
        for i in self.sensor_btns:
            i.setEnabled(False)
        for i, s in enumerate(self.sensors):
            # print(s.connected)
            if s.connected:
                self.sensor_btns[i].setVisible(True)
                self.sensor_frames[i].setVisible(True)
                self.sensor_btns[i].setEnabled(True)
                self.sensor_btns[i].setText(f'{self.sensors[i].name}')
                self.sensors[i].id = i
                self.sensors[i].connected = True
                self.sensor_btns_start[i].setEnabled(True)
                self.sensor_btns_stop[i].setEnabled(True)
                self.sensors[i].name = self.sensors[i].name
                self.sensor_names[i].setText(self.sensors[i].name)
            else:
                self.sensor_btns[i].setText(f'{i + 1} датчик')
                self.sensor_btns[i].setVisible(False)
                self.sensor_frames[i].setVisible(False)
                self.sensor_names[i].setText("Не подключен")
                self.sensor_values[i].setText("------")
                self.sensor_btns_start[i].setEnabled(False)
                self.sensor_btns_stop[i].setEnabled(False)
                self.sensors[i].id = i
                if self.timers[i].isActive():
                    self.timers[i].stop()
                    self.sensors[i].running = False

        # self.sensor_btns[8].setVisible(False)

        if True not in [i.running for i in self.sensors]:
            if self.dl.start_condition:
                self.dl.set_start_condition(0x21)
        '''for i in self.sensors_id:
            self.sensor_btns[i].setEnabled(True)
        for i in range(len(self.sensor_names)):
            if i not in self.sensors_id:
                self.sensor_names[i].setText("Не подключен")
                self.sensor_values[i].setText("------")
                self.sensor_btns_start[i].setEnabled(False)
                self.sensor_btns_stop[i].setEnabled(False)
                self.sensors[i].id = i
            else:
                self.sensors[i].id = i
                self.sensors[i].connected = True
                self.sensors[i].name = f'Датчик {i+1}'
                self.sensor_names[i].setText(self.sensors[i].name)'''

    def set_shadow(self):
        for x in shadow_elements:
            effect = QtWidgets.QGraphicsDropShadowEffect(self)
            effect.setBlurRadius(3)
            effect.setXOffset(0)
            effect.setYOffset(0)
            effect.setColor(QColor(0, 0, 0, 50))
            getattr(self.ui, x).setGraphicsEffect(effect)

    def closeEvent(self, event):
        self.dl.send_exit()


if __name__ == "__main__":
    if platform == "win32":
        multiprocessing.freeze_support()
    app = QApplication(sys.argv)

    window = MainWindow()
    window.resize(1300, 800)
    window.show()

    sys.exit(app.exec())
